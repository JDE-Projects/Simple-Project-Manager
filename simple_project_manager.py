"""
Simple Project Manager — an IT project planner/tracker.

JDE-Projects "Simple X Tool": Python 3 + PySide6/pywebview, single-file UI.
Phases -> Steps -> Actions (+ Contacts), with owners, deadlines, status,
priority, tags, notes and progress roll-up. The .xlsx IS the save format.

Phase 3b scope: per-item editing, .xlsx Save/Open/Save As (the workbook IS
the save file), native dialogs, and a silent recovery copy of unsaved work.
"""
import json
import os
import sys
import datetime
import threading
import time
import urllib.request
import ctypes
import ctypes.wintypes as wintypes

import webview
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

APP_VERSION = "1.3.0"
GITHUB_OWNER = "JDE-Projects"
GITHUB_REPO = "Simple-Project-Manager"

# Status / priority vocabularies — kept here so backend and UI agree.
STATUSES = ["Not started", "In progress", "Done", "Blocked"]
PRIORITIES = ["Low", "Normal", "High", "Critical"]

# .xlsx schema — one sheet, one row per item; Type drives the tree nesting.
SCHEMA_VERSION = "1"
COLUMNS = ["Type", "Name", "Owner", "Co-owner", "Deadline", "Status",
           "Progress", "Priority", "Tags", "Notes"]
LEVELS = {"Phase": 0, "Step": 1, "Action": 2, "Contact": 2}


def _to_date(value):
    """Coerce a stored value to a date for a real Excel date cell, else None."""
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    s = str(value or "").strip()
    if not s:
        return None
    try:
        return datetime.datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def resource_path(rel: str) -> str:
    """Path to a bundled resource, working both from source and PyInstaller."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, rel)


def app_dir() -> str:
    """Folder the app lives in — next to the .exe when frozen, else the script."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# --- Local prefs store (module-level) ------------------------------------
# One JSON file next to the app holds EVERY persisted setting: theme, window
# geometry, and anything added later. Always read-merge-write through
# load_prefs / save_prefs. Never overwrite the file with a single key, or the
# next setting you add silently wipes the others.

def _pref_path() -> str:
    return os.path.join(app_dir(), "simple_project_manager.pref")

def load_prefs() -> dict:
    """Load the full prefs dict. Tolerant of a missing or corrupt file."""
    try:
        with open(_pref_path(), "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def save_prefs(prefs: dict) -> bool:
    try:
        with open(_pref_path(), "w", encoding="utf-8") as f:
            json.dump(prefs, f)
        return True
    except Exception:
        return False


# --- Window geometry persistence ------------------------------------------
# Save and restore the ABSOLUTE window frame rectangle via Win32, found by the
# window title. GetWindowRect (save) and SetWindowPos (restore) share one
# frame-based, physical-pixel coordinate space, so the rect round-trips exactly
# at any DPI or monitor layout. Do NOT pass x/y into create_window and do NOT
# use window.move: pywebview's Qt backend applies those pre-show and relative to
# the primary screen, so the window lands on the wrong monitor, drifts down by
# the title-bar height each launch, and slides sideways at non-100% scaling.

def _win32():
    u = ctypes.windll.user32
    u.FindWindowW.restype = wintypes.HWND
    u.FindWindowW.argtypes = [wintypes.LPCWSTR, wintypes.LPCWSTR]
    u.GetWindowRect.argtypes = [wintypes.HWND, ctypes.POINTER(wintypes.RECT)]
    u.SetWindowPos.argtypes = [wintypes.HWND, wintypes.HWND, ctypes.c_int, ctypes.c_int,
                               ctypes.c_int, ctypes.c_int, wintypes.UINT]
    return u


def _save_geometry(win) -> None:
    """Save the absolute frame rect (physical px) via Win32. Wire to `closing`.
    Wrapped end to end so a failure here can never block the window from closing."""
    try:
        u = _win32()
        hwnd = u.FindWindowW(None, win.title)
        if not hwnd:
            return
        r = wintypes.RECT()
        if not u.GetWindowRect(hwnd, ctypes.byref(r)):
            return
        x, y, w, h = r.left, r.top, r.right - r.left, r.bottom - r.top
        if x <= -30000 or y <= -30000:   # minimized sentinel, not a real spot
            return
        if w <= 0 or h <= 0:
            return
        prefs = load_prefs()
        prefs["window"] = {"x": x, "y": y, "width": w, "height": h}
        save_prefs(prefs)
    except Exception:
        pass


def _restore_geometry(win) -> None:
    """Restore the saved frame rect via Win32. Wire to `shown` (after the OS
    window exists). Validate before applying; never raise."""
    try:
        geo = load_prefs().get("window")
        if not isinstance(geo, dict):
            return
        x, y, w, h = geo.get("x"), geo.get("y"), geo.get("width"), geo.get("height")
        for v in (x, y, w, h):
            if not isinstance(v, int) or isinstance(v, bool):
                return
        if w <= 0 or h <= 0:
            return
        # Is a point in the title bar still on a connected monitor?
        point = wintypes.POINT(x + 100, y + 30)
        user32 = ctypes.windll.user32
        user32.MonitorFromPoint.argtypes = [wintypes.POINT, wintypes.DWORD]
        user32.MonitorFromPoint.restype = wintypes.HMONITOR
        if not user32.MonitorFromPoint(point, 0):   # MONITOR_DEFAULTTONULL
            return
        u = _win32()
        hwnd = u.FindWindowW(None, win.title)
        if not hwnd:
            return
        SWP_NOZORDER, SWP_NOACTIVATE = 0x0004, 0x0010
        u.SetWindowPos(hwnd, None, x, y, w, h, SWP_NOZORDER | SWP_NOACTIVATE)
    except Exception:
        pass


class Api:
    """Bridge exposed to the UI. Methods return JSON-able values; the UI awaits."""

    def __init__(self):
        self._window = None
        self._debug = False
        self._debug_path = None
        self._filename = ""

    def set_window(self, w):
        self._window = w

    # --- document data (stubbed in 3a; real .xlsx I/O in 3b) ---------------
    def get_state(self):
        """Initial payload the UI loads on startup."""
        return {
            "version": APP_VERSION,
            "theme": self._load_theme(),
            "statuses": STATUSES,
            "priorities": PRIORITIES,
            "items": [],
            "filename": "",
        }

    # --- progress roll-up (auto column; UI shows it richer in 3c) ----------
    def _progress_map(self, items):
        """id -> percent for phases/steps. Actions render as ✔/☐ instead."""
        prog, structure, phase, step = {}, [], None, None
        for it in items:
            t = it.get("type")
            if t == "Phase":
                phase = {"id": it["id"], "steps": []}
                step = None
                structure.append(phase)
            elif t == "Step":
                step = {"id": it["id"], "status": it.get("status", ""), "actions": []}
                if phase is None:
                    phase = {"id": None, "steps": []}
                    structure.append(phase)
                phase["steps"].append(step)
            elif t == "Action" and step is not None:
                step["actions"].append(it.get("status", ""))
        for ph in structure:
            pcts = []
            for st in ph["steps"]:
                if st["actions"]:
                    done = sum(1 for s in st["actions"] if s == "Done")
                    pct = round(100 * done / len(st["actions"]))
                else:
                    pct = 100 if st["status"] == "Done" else 0
                prog[st["id"]] = pct
                pcts.append(pct)
            if ph["id"] is not None:
                prog[ph["id"]] = round(sum(pcts) / len(pcts)) if pcts else 0
        return prog

    # --- .xlsx writer / reader (the workbook IS the save file) -------------
    def _write_xlsx(self, path, items, source=""):
        wb = Workbook()
        ws = wb.active
        ws.title = "Project"
        ws.append(COLUMNS)
        head_fill = PatternFill("solid", fgColor="178577")
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

        prog = self._progress_map(items)
        r = 1
        for it in items:
            r += 1
            typ = it.get("type", "")
            lvl = LEVELS.get(typ, 0)
            name_cell = ws.cell(r, 2, it.get("name", ""))
            name_cell.alignment = Alignment(indent=lvl, vertical="center")
            if typ == "Phase":
                name_cell.font = Font(bold=True)
            ws.cell(r, 1, typ)
            ws.cell(r, 3, it.get("owner", ""))
            ws.cell(r, 4, it.get("coowner", ""))
            d = _to_date(it.get("deadline", ""))
            dcell = ws.cell(r, 5)
            if d is not None:
                dcell.value = d
                dcell.number_format = "yyyy-mm-dd"
            else:
                dcell.value = it.get("deadline", "")
            ws.cell(r, 6, it.get("status", ""))
            pcell = ws.cell(r, 7)
            if typ == "Action":
                pcell.value = "✔" if it.get("status") == "Done" else "☐"
                pcell.alignment = Alignment(horizontal="center")
            elif it.get("id") in prog:
                pcell.value = prog[it["id"]] / 100.0
                pcell.number_format = "0%"
            ws.cell(r, 8, it.get("priority", "") if typ in ("Step", "Action") else "")
            ws.cell(r, 9, ",".join(it.get("tags") or []))
            ws.cell(r, 10, it.get("notes", ""))
            if lvl:
                ws.row_dimensions[r].outline_level = lvl
        ws.sheet_properties.outlinePr.summaryBelow = False

        for i, w in enumerate([9, 42, 12, 12, 13, 13, 10, 10, 20, 44], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        last = max(r, 2)
        dv_s = DataValidation(type="list",
                              formula1='"%s"' % ",".join(STATUSES), allow_blank=True)
        ws.add_data_validation(dv_s)
        dv_s.add(f"F2:F{last}")
        dv_p = DataValidation(type="list",
                              formula1='"%s"' % ",".join(PRIORITIES), allow_blank=True)
        ws.add_data_validation(dv_p)
        dv_p.add(f"H2:H{last}")

        meta = wb.create_sheet("_spm")
        meta.sheet_state = "hidden"
        meta["A1"], meta["B1"] = "app", "Simple Project Manager"
        meta["A2"], meta["B2"] = "schema", SCHEMA_VERSION
        meta["A3"], meta["B3"] = "tag_colors", json.dumps({})
        meta["A4"], meta["B4"] = "source", source
        wb.save(path)

    def _read_xlsx(self, path):
        wb = load_workbook(path, data_only=True)
        is_ours = "_spm" in wb.sheetnames
        ws = wb["Project"] if "Project" in wb.sheetnames else wb.active
        header = [str(c.value).strip().lower() if c.value is not None else ""
                  for c in ws[1]]
        idx = {h: i for i, h in enumerate(header)}

        def get(row, *names, default=""):
            for n in names:
                i = idx.get(n)
                if i is not None and i < len(row) and row[i] is not None:
                    return row[i]
            return default

        items, n = [], 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            typ = str(get(row, "type") or "").strip()
            if typ not in ("Phase", "Step", "Action", "Contact"):
                continue
            n += 1
            dl = get(row, "deadline")
            if isinstance(dl, (datetime.datetime, datetime.date)):
                dl = dl.strftime("%Y-%m-%d")
            else:
                dl = str(dl).strip() if dl else ""
            raw_tags = str(get(row, "tags") or "")
            tags = [t.strip() for t in raw_tags.split(",") if t.strip()]
            items.append({
                "id": f"i{n}", "type": typ,
                "name": str(get(row, "name") or "").strip(),
                "owner": str(get(row, "owner") or "").strip(),
                "coowner": str(get(row, "co-owner", "coowner") or "").strip(),
                "deadline": dl,
                "status": str(get(row, "status") or "").strip(),
                "priority": (str(get(row, "priority") or "").strip() or "Normal"),
                "tags": tags,
                "notes": str(get(row, "notes") or "").strip(),
            })
        warn = "" if is_ours else ("Opened best-effort — this workbook has no "
                                   "Simple Project Manager marker.")
        return items, warn

    # --- document commands (called from the UI) ----------------------------
    def new_project(self):
        self._filename = ""
        self._clear_recovery()
        self.log("New project")
        return {"ok": True, "items": [], "filename": ""}

    def open_project(self):
        try:
            dlg = webview.FileDialog.OPEN
        except AttributeError:  # older pywebview
            dlg = webview.OPEN_DIALOG
        try:
            sel = self._window.create_file_dialog(
                dlg, allow_multiple=False,
                file_types=("Excel workbook (*.xlsx)", "All files (*.*)"))
        except Exception as e:
            self.log(f"Open dialog failed: {e}")
            return {"ok": False, "error": "Couldn't open the file picker."}
        if not sel:
            return {"cancelled": True}
        path = sel[0] if isinstance(sel, (list, tuple)) else sel
        try:
            items, warn = self._read_xlsx(path)
        except Exception as e:
            self.log(f"Open failed for {path}: {e}")
            return {"ok": False,
                    "error": "That file couldn't be opened as a project workbook."}
        self._filename = path
        self._clear_recovery()
        self.log(f"Opened {path} ({len(items)} rows)")
        return {"ok": True, "items": items, "filename": path, "warn": warn}

    def save_project(self, items):
        if not self._filename:
            return self.save_as(items)
        return self._do_save(self._filename, items)

    def save_as(self, items):
        suggested = os.path.basename(self._filename) if self._filename else "project.xlsx"
        try:
            dlg = webview.FileDialog.SAVE
        except AttributeError:  # older pywebview
            dlg = webview.SAVE_DIALOG
        try:
            sel = self._window.create_file_dialog(
                dlg, save_filename=suggested,
                file_types=("Excel workbook (*.xlsx)",))
        except Exception as e:
            self.log(f"Save dialog failed: {e}")
            return {"ok": False, "error": "Couldn't open the save dialog."}
        if not sel:
            return {"cancelled": True}
        path = sel[0] if isinstance(sel, (list, tuple)) else sel
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        return self._do_save(path, items)

    def _do_save(self, path, items):
        try:
            self._write_xlsx(path, items)
        except PermissionError:
            self.log(f"Save denied (file open in Excel?): {path}")
            return {"ok": False,
                    "error": "Couldn't save — is the file open in Excel?"}
        except Exception as e:
            self.log(f"Save failed for {path}: {e}")
            return {"ok": False, "error": "Couldn't save the project file."}
        self._filename = path
        self._clear_recovery()
        self.log(f"Saved {path}")
        return {"ok": True, "filename": path}

    # --- silent recovery copy of unsaved work ------------------------------
    def _recovery_path(self):
        return os.path.join(app_dir(), ".spm_recovery.xlsx")

    def write_recovery(self, items):
        try:
            self._write_xlsx(self._recovery_path(), items, source=self._filename)
        except Exception as e:
            self.log(f"Recovery write failed: {e}")
            return {"ok": False}
        return {"ok": True}

    def _clear_recovery(self):
        try:
            p = self._recovery_path()
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            pass

    def clear_recovery(self):
        self._clear_recovery()
        return {"ok": True}

    def check_recovery(self):
        p = self._recovery_path()
        if not os.path.exists(p):
            return {"exists": False}
        try:
            items, _ = self._read_xlsx(p)
            wb = load_workbook(p, data_only=True)
            src = ""
            if "_spm" in wb.sheetnames:
                src = str(wb["_spm"]["B4"].value or "")
        except Exception as e:
            self.log(f"Recovery read failed: {e}")
            self._clear_recovery()
            return {"exists": False}
        self._filename = src
        return {"exists": True, "items": items, "filename": src}

    # --- theme preference (through the shared prefs store) ------------------
    def _load_theme(self) -> str:
        theme = load_prefs().get("theme")
        return theme if theme in ("dark", "light") else "dark"

    def save_theme(self, theme: str):
        if theme not in ("dark", "light"):
            return {"ok": False}
        prefs = load_prefs()
        prefs["theme"] = theme
        if save_prefs(prefs):
            self.log(f"Theme set to {theme}")
            return {"ok": True}
        self.log("Could not save theme pref: save_prefs failed")
        return {"ok": False}

    # --- misc bridge helpers ------------------------------------------------
    def open_url(self, url: str):
        """Open a link in the system browser — never navigate the app window."""
        import webbrowser
        webbrowser.open(url)
        return {"ok": True}

    def check_update(self):
        """Compare the latest published release to APP_VERSION. Silent on failure."""
        result = {"version": APP_VERSION, "latest": None, "update": False, "url": ""}
        try:
            url = (f"https://api.github.com/repos/{GITHUB_OWNER}/"
                   f"{GITHUB_REPO}/releases/latest")
            req = urllib.request.Request(url, headers={"Accept": "application/vnd.github+json"})
            with urllib.request.urlopen(req, timeout=4) as r:
                data = json.load(r)
            latest = (data.get("tag_name") or "").lstrip("v")
            result["latest"] = latest
            result["url"] = data.get("html_url", "")
            if latest and self._is_newer(latest, APP_VERSION):
                result["update"] = True
        except Exception:
            pass  # offline / private repo / rate-limited — stay quiet
        return result

    @staticmethod
    def _is_newer(latest: str, current: str) -> bool:
        def parts(v):
            out = []
            for p in v.split("."):
                try:
                    out.append(int(p))
                except ValueError:
                    out.append(0)
            return out
        return parts(latest) > parts(current)

    # --- debug log ----------------------------------------------------------
    def set_debug(self, on: bool):
        self._debug = bool(on)
        if self._debug and not self._debug_path:
            stamp = datetime.datetime.now().strftime("%m%d%Y_%H%M%S")
            self._debug_path = os.path.join(app_dir(), f"Debug_Log_{stamp}.txt")
            self.log("Debug log started")
        return {"ok": True}

    def log(self, msg: str):
        if not self._debug or not self._debug_path:
            return
        try:
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(self._debug_path, "a", encoding="utf-8") as f:
                f.write(f"[{ts}] {msg}\n")
        except Exception:
            pass


# Splash close: honor a 5s minimum so it doesn't just flash, but never
# hang past 30s. Whichever of (window ready after the floor) / (watchdog)
# fires first wins; the rest are no-ops. In source/dev runs pyi_splash is
# absent, so all of this does nothing.
_splash = {"closed": False, "start": time.monotonic()}

def _close_splash():
    if _splash["closed"]:
        return
    _splash["closed"] = True
    try:
        import pyi_splash  # only present in the frozen build
        pyi_splash.close()
    except Exception:
        pass

def _on_window_ready():
    elapsed = time.monotonic() - _splash["start"]
    if elapsed >= 5:
        _close_splash()
    else:
        threading.Timer(5 - elapsed, _close_splash).start()


_mutex_handle = None   # module-level: must live for the process lifetime

def _acquire_single_instance(mutex_name: str) -> bool:
    # Name convention: "JDE_Simple{Thing}Tool_SingleInstance"
    # Session-local (no "Global\" prefix): each Windows session (e.g. RDP,
    # fast user switching) gets its own instance instead of colliding across users.
    global _mutex_handle
    try:
        # use_last_error=True: ctypes.windll's GetLastError() can be clobbered
        # by ctypes-internal calls, so read the error via ctypes.get_last_error() instead.
        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        _mutex_handle = kernel32.CreateMutexW(None, False, mutex_name)
        return ctypes.get_last_error() != 183   # ERROR_ALREADY_EXISTS
    except Exception:
        return True   # fail open: never block launch over a mutex error

def _focus_existing_window(app_title: str) -> bool:
    # Best-effort only: any failure here must not stop the caller from deciding what to do next.
    try:
        user32 = ctypes.windll.user32
        found = {"hwnd": None}

        WNDENUMPROC = ctypes.WINFUNCTYPE(wintypes.BOOL, wintypes.HWND, wintypes.LPARAM)

        def _enum_proc(hwnd, lparam):
            if not user32.IsWindowVisible(hwnd):
                return True
            length = user32.GetWindowTextLengthW(hwnd)
            if length == 0:
                return True
            buf = ctypes.create_unicode_buffer(length + 1)
            user32.GetWindowTextW(hwnd, buf, length + 1)
            # Exact match only: a prefix match could hit an unrelated window
            # (e.g. a browser tab starting with the app name). A miss falls
            # through to a normal launch anyway.
            if buf.value == app_title:
                found["hwnd"] = hwnd
                return False   # stop enumerating, match found
            return True

        user32.EnumWindows(WNDENUMPROC(_enum_proc), 0)

        hwnd = found["hwnd"]
        if not hwnd:
            return False

        SW_RESTORE = 9
        if user32.IsIconic(hwnd):
            user32.ShowWindow(hwnd, SW_RESTORE)
        user32.SetForegroundWindow(hwnd)
        return True
    except Exception:
        return False

def main():
    if not _acquire_single_instance("JDE_SimpleProjectManager_SingleInstance"):
        if _focus_existing_window("Simple Project Manager"):
            sys.exit(0)
        # Existing window not found: fail open and launch normally.

    if sys.platform == "win32":
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(
            "JDEProjects.SimpleProjectManager"
        )

    api = Api()
    win = webview.create_window(
        "Simple Project Manager",
        url=resource_path("simple_project_manager-UI.html"),
        js_api=api,
        width=1180, height=820, min_size=(900, 600),
        background_color="#0a0e14",
    )
    api.set_window(win)

    win.events.shown += lambda: _restore_geometry(win)

    def _on_closing():
        _save_geometry(win)
        return True
    win.events.closing += _on_closing

    win.events.loaded += _on_window_ready
    threading.Timer(30, _close_splash).start()  # ceiling: never hang
    try:
        webview.start(gui="qt", icon=resource_path("simple_project_manager.png"))
    except TypeError:
        webview.start(gui="qt")


if __name__ == "__main__":
    main()
